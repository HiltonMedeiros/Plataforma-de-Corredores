from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required, user_passes_test
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
from .models import User, Atividade, Inscricao, Bairro, ValidacaoResidencia

def is_admin(user):
    return user.is_staff

# ==================== EXPORTAR CSV ====================

@require_http_methods(["GET"])
@login_required
@user_passes_test(is_admin)
def exportar_inscritos_csv(request):
    """Exporta lista de inscritos em formato CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inscritos.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Usuário', 'Email', 'CPF', 'Bairro', 'Etapa', 'Data Inscrição', 'Status'])

    inscricoes = Inscricao.objects.select_related(
        'user', 'etapa', 'user__perfil', 'user__perfil__bairro'
    ).all()

    for inscricao in inscricoes:
        writer.writerow([
            inscricao.id,
            inscricao.user.username,
            inscricao.user.email,
            inscricao.user.perfil.cpf,
            inscricao.user.perfil.bairro.nome,
            inscricao.etapa.nome,
            inscricao.data_inscricao.strftime('%d/%m/%Y'),
            inscricao.status,
        ])

    return response


@require_http_methods(["GET"])
@login_required
@user_passes_test(is_admin)
def exportar_atividades_csv(request):
    """Exporta atividades aprovadas em formato CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="atividades_aprovadas.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Usuário', 'Email', 'Etapa', 'Distância (km)', 'Tempo',
        'Data', 'Tipo', 'Auditado Por', 'Data Validação', 'Status'
    ])

    atividades = Atividade.objects.filter(
        status_validacao='APROVADO'
    ).select_related(
        'user', 'etapa', 'auditado_por'
    )

    for atividade in atividades:
        writer.writerow([
            atividade.id,
            atividade.user.username,
            atividade.user.email,
            atividade.etapa.nome,
            atividade.distancia,
            atividade.tempo_total,
            atividade.data_atividade.strftime('%d/%m/%Y'),
            atividade.tipo_registro,
            atividade.auditado_por.username if atividade.auditado_por else '-',
            atividade.data_validacao.strftime('%d/%m/%Y %H:%M') if atividade.data_validacao else '-',
            atividade.status_validacao,
        ])

    return response


@require_http_methods(["GET"])
@login_required
@user_passes_test(is_admin)
def exportar_por_bairro_csv(request):
    """Exporta estatísticas agrupadas por bairro em formato CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="estatisticas_bairros.csv"'

    writer = csv.writer(response)
    writer.writerow(['Bairro', 'Total Inscritos', 'Total Atividades Aprovadas', 'Distância Total (km)'])

    bairros = Bairro.objects.all()

    for bairro in bairros:
        inscritos = Inscricao.objects.filter(
            user__perfil__bairro=bairro,
            status='ATIVA'
        ).count()

        atividades = Atividade.objects.filter(
            user__perfil__bairro=bairro,
            status_validacao='APROVADO'
        )

        total_atividades = atividades.count()
        distancia_total = sum(float(a.distancia) for a in atividades)

        writer.writerow([
            bairro.nome,
            inscritos,
            total_atividades,
            f"{distancia_total:.2f}",
        ])

    return response


# ==================== EXPORTAR PDF ====================

@require_http_methods(["GET"])
@login_required
@user_passes_test(is_admin)
def exportar_relatorio_pdf(request):
    """Exporta relatório completo em PDF."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_bayeux.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1,
    )
    title = Paragraph("📊 Relatório Bayeux Movimenta", title_style)
    story.append(title)

    # Data do Relatório
    date_text = Paragraph(
        f"<b>Data do Relatório:</b> {datetime.now().strftime('%d de %B de %Y às %H:%M')}",
        styles['Normal']
    )
    story.append(date_text)
    story.append(Spacer(1, 0.3 * inch))

    # Estatísticas Gerais
    story.append(Paragraph("<b>1. Estatísticas Gerais</b>", styles['Heading2']))

    total_usuarios = User.objects.count()
    total_inscritos = Inscricao.objects.filter(status='ATIVA').count()
    total_atividades = Atividade.objects.filter(status_validacao='APROVADO').count()
    validacoes_pendentes = ValidacaoResidencia.objects.filter(status='PENDENTE').count()

    stats_data = [
        ['Métrica', 'Quantidade'],
        ['Total de Usuários', str(total_usuarios)],
        ['Total de Inscritos Ativos', str(total_inscritos)],
        ['Atividades Aprovadas', str(total_atividades)],
        ['Validações Pendentes', str(validacoes_pendentes)],
    ]

    stats_table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.3 * inch))

    # Estatísticas por Bairro
    story.append(Paragraph("<b>2. Distribuição por Bairro</b>", styles['Heading2']))

    bairro_data = [['Bairro', 'Inscritos', 'Atividades Aprovadas']]
    bairros = Bairro.objects.all()

    for bairro in bairros:
        inscritos = Inscricao.objects.filter(
            user__perfil__bairro=bairro,
            status='ATIVA'
        ).count()

        atividades = Atividade.objects.filter(
            user__perfil__bairro=bairro,
            status_validacao='APROVADO'
        ).count()

        bairro_data.append([bairro.nome, str(inscritos), str(atividades)])

    bairro_table = Table(bairro_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
    bairro_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#764ba2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(bairro_table)

    doc.build(story)
    return response


# ==================== EXPORTAR EXCEL ====================

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    @require_http_methods(["GET"])
    @login_required
    @user_passes_test(is_admin)
    def exportar_completo_excel(request):
        """Exporta dados completos em Excel com múltiplas abas."""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        # Aba 1: Inscritos
        ws_inscritos = wb.create_sheet("Inscritos")
        headers = ['ID', 'Usuário', 'Email', 'CPF', 'Bairro', 'Etapa', 'Data Inscrição', 'Status']
        ws_inscritos.append(headers)

        inscricoes = Inscricao.objects.select_related(
            'user', 'etapa', 'user__perfil', 'user__perfil__bairro'
        )

        for inscricao in inscricoes:
            ws_inscritos.append([
                inscricao.id,
                inscricao.user.username,
                inscricao.user.email,
                inscricao.user.perfil.cpf,
                inscricao.user.perfil.bairro.nome,
                inscricao.etapa.nome,
                inscricao.data_inscricao.strftime('%d/%m/%Y'),
                inscricao.status,
            ])

        # Aba 2: Atividades
        ws_atividades = wb.create_sheet("Atividades Aprovadas")
        headers = ['ID', 'Usuário', 'Etapa', 'Distância (km)', 'Data', 'Status']
        ws_atividades.append(headers)

        atividades = Atividade.objects.filter(
            status_validacao='APROVADO'
        ).select_related('user', 'etapa')

        for atividade in atividades:
            ws_atividades.append([
                atividade.id,
                atividade.user.username,
                atividade.etapa.nome,
                float(atividade.distancia),
                atividade.data_atividade.strftime('%d/%m/%Y'),
                atividade.status_validacao,
            ])

        # Aba 3: Por Bairro
        ws_bairro = wb.create_sheet("Por Bairro")
        headers = ['Bairro', 'Inscritos', 'Atividades Aprovadas', 'Distância Total']
        ws_bairro.append(headers)

        for bairro in Bairro.objects.all():
            inscritos = Inscricao.objects.filter(
                user__perfil__bairro=bairro,
                status='ATIVA'
            ).count()

            atividades_obj = Atividade.objects.filter(
                user__perfil__bairro=bairro,
                status_validacao='APROVADO'
            )

            total_atividades = atividades_obj.count()
            distancia_total = sum(float(a.distancia) for a in atividades_obj)

            ws_bairro.append([bairro.nome, inscritos, total_atividades, distancia_total])

        # Configurar largura das colunas
        for ws in [ws_inscritos, ws_atividades, ws_bairro]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_bayeux.xlsx"'
        wb.save(response)
        return response

except ImportError:
    pass  # openpyxl não instalado
